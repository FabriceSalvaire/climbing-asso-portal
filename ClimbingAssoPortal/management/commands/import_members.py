####################################################################################################
#
# Climbing Asso Portal - A Portal for Climbing Club (Association)
# Copyright (C) 2018 Fabrice Salvaire
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.
#
####################################################################################################

####################################################################################################
#
# cf. https://docs.djangoproject.com/en/2.0/howto/custom-management-commands/
# https://docs.djangoproject.com/en/2.0/topics/auth/default/
#
####################################################################################################

####################################################################################################

import datetime

from django.core.management.base import BaseCommand, CommandError

from django.contrib.auth.models import User, Group
# from account.conf import settings
# settings.AUTH_USER_MODEL = auth.User

from ClimbingAssoPortal.models.city import FrenchCity
from ClimbingAssoPortal.models.member import Club, Member, ClubMember
from ClimbingAssoPortal.tools.TextTools import strip_accents

from .MemberFixes import DOMAIN_FIXES, CLUB_FIXES

####################################################################################################

class Command(BaseCommand):

    help = 'Import members'

    ##############################################

    def _log_success(self, message):
        self.stdout.write(self.style.SUCCESS(message))

    ##############################################

    def _log_warning(self, message):
        self.stdout.write(self.style.WARNING(message))

    ##############################################

    def _log_error(self, message):
        self.stdout.write(self.style.ERROR(message))

    ##############################################

    def add_arguments(self, parser):

        parser.add_argument(
            'input',
            metavar='InputFile',
            help='Input file',
        )

    ##############################################

    def handle(self, *args, **options):

        input_file = options['input']

        self._drop_members()
        self._fetch_data()

        if input_file.endswith('.xlsx'):
            self._read_xlsls(input_file)

        # try:
        #     pass
        # except Exception:
        #     raise CommandError()

        self._log_success('Success')

    ##############################################

    def _fetch_data(self):

        self._member_group = Group.objects.get(name='Membre')
        self._clubs = {club.name:club for club in Club.objects.all()}

    ##############################################

    def _drop_members(self):

        User.objects.filter(is_staff=False).delete()
        Member.objects.filter().delete()
        ClubMember.objects.filter().delete()

    ##############################################

    def _read_xlsls(self, input_file):

        self._email_domains = set()

        from openpyxl import load_workbook
        work_book = load_workbook(filename=input_file, read_only=True)
        for sheet_name in work_book.sheetnames:
            sheet = work_book[sheet_name]
            for row in sheet.rows:
                self._process_row(row)

        # for domain in sorted(self._email_domains):
        #     print(domain)

    ##############################################

    @staticmethod
    def _clean_name(name):

        if name.endswith('.'):
            name = name[:-1]
        return name

    ##############################################

    @staticmethod
    def _fix_firstname(first_name):

        for splitter in (' ', '-'):
            parts = [part.strip() for part in first_name.split(splitter)]
            first_name = splitter.join([part.title() for part in parts if part])
        return first_name

    ##############################################

    @staticmethod
    def _fixe_phone(number):

        if number is None:
            return None

        number = str(number)
        if number.endswith('(?)'): # Fixme: ???
            number = number[:-3]

        try:
            int(number)
        except:
            raise NameError('Phone number is not a int {}'.format(number))

        if number.startswith('33') and len(number) == 11:
            number = '0' + number[2:]
        elif not number.startswith('0') and len(number) == 9:
            number = '0' + number

        if len(number) != 10:
            raise NameError('Invalid phone number length {}'.format(number))

        return number

    ##############################################

    def _fix_email(self, email):

        if not email:
            return None

        name, domain = email.split('@')

        self._email_domains.add(domain)
        domain_fix = DOMAIN_FIXES.get(domain, None)
        if domain_fix is not None:
            domain = domain_fix

        name_parts = name.split('.')
        domains_parts = domain.split('.')
        parts = name_parts + domains_parts
        title_hits = 0
        for part in parts:
            if part.istitle():
                title_hits += 1
        if title_hits == len(parts):
            name = name.lower()
            domain = domain.lower()

        return '{}@{}'.format(name, domain)

    ##############################################

    @staticmethod
    def _retrieve_city(zip_code, city):

        city_obj = None

        if zip_code is not None:
            results = FrenchCity.objects.filter(zip_code=zip_code)
            if len(results) == 1:
                city_obj = results[0]

        if city_obj is None and city is not None:
            results = FrenchCity.objects.filter(name=city.upper())
            if len(results) == 1:
                city_obj = results[0]

        return city_obj

    ##############################################

    def _process_row(self, row):

        # 'IdMembre', "ANNÉED'INSCRIPTION", 'GROUPE 2016/17', 'n° licence FSGT', 'Club de la licence',
        # 'NOM', 'PRENOM', 'DTNAISS', 'SEXE',
        # 'ADRESSE', 'CODEPOSTAL', 'VILLE',
        # 'TELDOM', 'TELPRO', 'TELPORT', 'EMAIL',
        # 'Nom du saisisseur',
        # 'TARIF 2017/18',
        # 'Date du certificat médical',
        # 'PÔLE1', 'PÔLE2',
        # 'No Chèque', 'Nom Titulaire du chèque', 'Nom Banque', 'Montant total du chèque',
        # 'Commentaire'

        values = []
        for cell in row:
            value = cell.value
            if isinstance(value, str):
                value = value.strip()
            values.append(value)

        (
            membre_id,
            registration_year,
            group,
            license_id, license_club,
            last_name, first_name,
            birth_date, sex,
            address, zip_code, city,
            phone_home, phone_work, phone_mobile,
            email,
            operator_name,
            adhesion_price,
            medical_certificate_year,
            pole1, pole2,
            check_id, check_name, check_bank, check_amount,
            comment
        ) = values

        if last_name == 'NOM' or last_name is None:
            return

        last_name = self._clean_name(last_name).upper()
        first_name = self._fix_firstname(self._clean_name(first_name))
        sex = sex.lower()
        birth_date = datetime.date(birth_date.year, birth_date.month, birth_date.day)

        city_obj = self._retrieve_city(zip_code, city)
        if city_obj is None:
            self._log_warning('Error: cannot found city for {} {} : {} {}'.format(last_name, first_name, zip_code, city))

        email = self._fix_email(email)
        phone_mobile = self._fixe_phone(phone_mobile)
        phone_home = self._fixe_phone(phone_home)
        phone_work = self._fixe_phone(phone_work)

        license_id = int(license_id)
        license_club_obj = self._clubs[CLUB_FIXES[license_club]]

        group = 's' if group == 'SOIR' else 'm'
        if isinstance(adhesion_price, str):
            group_name = 'soir' if group == 's' else 'midi'
            adhesion_price = adhesion_price.lower()
            adhesion_price = adhesion_price.replace(group_name, '').strip()
            try:
                adhesion_price = int(adhesion_price)
            except:
                # Fixme: adhesion_price = '???'
                adhesion_price = 1000
        if ((group == 'm' and adhesion_price < 60) or
            (group == 's' and adhesion_price < 90)):
            social_discount = True
        else:
            social_discount = False

        username = strip_accents('{}-{}'.format(last_name, first_name).lower())
        username = username.replace(' ', '-')
        username = username.replace("'", '')

        sorted_fields = (
            last_name, first_name,
            birth_date, sex,
            license_id, license_club, group,
            email,
            phone_home, phone_work, phone_mobile,
        )

        # if not username.startswith('s'):
        #     return

        # print(' | '.join([str(x) for x in sorted_fields]))

        password = '{}{:02}'.format(license_id, birth_date.day)
        # print(' '*2, username, password, last_name, first_name)

        if True:
            if User.objects.filter(username=username):
                self._log_warning('Error: username {} clash for {} {}'.format(username, last_name, first_name))
            elif User.objects.filter(email=email):
                self._log_warning('Error: email {} clash for {} {}'.format(email, last_name, first_name))
            else:
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=password, # stored has hash
                    first_name=first_name,
                    last_name=last_name,
                )
                # user is now saved

                user.groups.set([self._member_group])
                user.save()

                account = user.account
                account.language = 'fr'
                account.timezone = 'Europe/Paris'
                account.save()

                member = Member(
                    user=user,
                    address=address,
                    city=city_obj,
                    license_club=license_club_obj,
                    license_id=license_id,
                    birth_date=birth_date,
                    sex=sex,
                    phone_home=phone_home,
                    phone_work=phone_work,
                    phone_mobile=phone_mobile,
                    medical_certificate_year=2017,
                )
                member.save()

                club_member = ClubMember(
                    member=member,
                    registration_year=registration_year,
                    group=group,
                    social_discount=social_discount,
                )
                club_member.save()
